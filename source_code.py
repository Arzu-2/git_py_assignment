import openpyxl
import matplotlib.pyplot as plt
import csv

# Function to analyze Log2FC values from the dataset
def analyze_log2fc(file_name, output_file):
    # Load the Excel file using openpyxl
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active  # Get the active sheet

    # Initialize lists to store positive and negative Log2FC values
    positive_log2fc = []
    negative_log2fc = []
    
    # Initialize counters for BM_unique YES and NO
    count_bm_yes = 0
    count_bm_no = 0

    # Loop through the rows and collect Log2FC values (assuming Log2FC is in the second column)
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):  # Start from row 2 to skip the header
        log2fc_value = row[0].value
        if log2fc_value is not None:
            if log2fc_value > 0:
                positive_log2fc.append(log2fc_value)
            elif log2fc_value < 0:
                negative_log2fc.append(log2fc_value)

    # Loop through the rows to count YES/NO in the 10th column (BM_unique column)
    for row in sheet.iter_rows(min_row=2, min_col=10, max_col=10):  # 10th column for BM_unique
        bm_value = row[0].value
        if bm_value == "YES":
            count_bm_yes += 1
        elif bm_value == "NO":
            count_bm_no += 1

    # Count the number of positive and negative Log2FC values
    num_positive = len(positive_log2fc)
    num_negative = len(negative_log2fc)

    # Calculate the average of positive and negative Log2FC values
    avg_positive_log2fc = sum(positive_log2fc) / num_positive if num_positive > 0 else 0
    avg_negative_log2fc = sum(negative_log2fc) / num_negative if num_negative > 0 else 0

    # Output the results
    results = (
        f"Number of positive Log2FC values: {num_positive}\n"
        f"Number of negative Log2FC values: {num_negative}\n"
        f"Average of positive Log2FC values: {avg_positive_log2fc}\n"
        f"Average of negative Log2FC values: {avg_negative_log2fc}\n"
        f"Number of BM_unique 'YES': {count_bm_yes}\n"
        f"Number of BM_unique 'NO': {count_bm_no}\n"
    )

    # Save the results to a CSV file
    with open(output_file, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Statistic", "Value"])
        writer.writerow(["Number of positive Log2FC values", num_positive])
        writer.writerow(["Number of negative Log2FC values", num_negative])
        writer.writerow(["Average of positive Log2FC values", avg_positive_log2fc])
        writer.writerow(["Average of negative Log2FC values", avg_negative_log2fc])
        writer.writerow(["Number of BM_unique 'YES'", count_bm_yes])
        writer.writerow(["Number of BM_unique 'NO'", count_bm_no])

    # Plot the distribution of positive and negative Log2FC values
    plt.figure(figsize=(8, 6))
    plt.hist(positive_log2fc, bins=20, alpha=0.7, label="Positive Log2FC", color='blue')
    plt.hist(negative_log2fc, bins=20, alpha=0.7, label="Negative Log2FC", color='red')
    plt.axvline(0, color='black', linestyle='dashed', linewidth=2)
    plt.xlabel('Log2FC')
    plt.ylabel('Frequency')
    plt.title('Distribution of Positive and Negative Log2FC Values')
    plt.legend()

    # Save the plot as an image file (e.g., PNG)
    plt.savefig("log2fc_histogram.png")
    plt.show()

# Main function to run the analysis
def main():
    # Set the file name (make sure this path is correct)
    file_name = 'DataSet_ALL_DE_proteins.xlsx'
    output_file = 'analysis_output.csv'  # Specify the output file for the statistics

    # Call the function to analyze Log2FC values
    analyze_log2fc(file_name, output_file)

# Run the main function
if __name__ == "__main__":
    main()